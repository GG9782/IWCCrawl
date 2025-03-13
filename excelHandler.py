import openpyxl
from openpyxl.utils import get_column_letter


def append_data_to_excel_with_format(file_path, sheet_name, data):
    """
    将数据追加到Excel文件的指定sheet的表格下方，同时保持原数据的格式不变。

    参数:
    file_path (str): Excel文件路径。
    sheet_name (str): 要操作的sheet名称。
    data (list of dict): 输入数据，每个元素为字典，key对应Excel里的字段名。

    返回:
    None
    """
    try:
        # 打开Excel文件并获取指定的sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
        return
    except KeyError:
        print(f"Sheet {sheet_name} 不存在。")
        return
    except Exception as e:
        print(f"打开Excel文件时出错: {e}")
        return

    # 获取Excel文件中第一行的字段名（即列名）
    existing_field_names = [cell.value for cell in sheet[1]]

    # 检查输入数据的字典key是否与Excel文件的字段名匹配
    # 如果不匹配，则打印警告并跳过不匹配的数据（这里可以根据需求调整处理方式）
    for item in data:
        for key in item.keys():
            if key not in existing_field_names:
                print(f"警告: 输入数据的key '{key}' 与Excel文件的字段名不匹配，将跳过此key的值。")

    # 找到Excel文件中最后一行的行号
    last_row = sheet.max_row

    # 遍历输入数据，并逐行追加到Excel文件的下方
    for row_data in data:
        new_row = []
        for col_num, field_name in enumerate(existing_field_names, start=1):
            cell_value = row_data.get(field_name)
            new_cell = sheet.cell(row=last_row + 1, column=col_num, value=cell_value)
            # 尝试复制原单元格的格式到新单元格（这里只复制了部分基本格式）
            try:
                original_cell = sheet.cell(row=1, column=col_num)  # 假设第一行的格式是我们要复制的
                new_cell.font = original_cell.font.copy()
                new_cell.fill = original_cell.fill.copy()
                new_cell.border = original_cell.border.copy()
                new_cell.alignment = original_cell.alignment.copy()
                new_cell.number_format = original_cell.number_format
            except Exception as e:
                print(f"复制单元格格式时出错: {e}")
        new_row.append(new_cell)  # 虽然这里append了，但实际上不需要，因为我们已经直接操作了sheet
        last_row += 1  # 更新最后一行的行号

    # 保存更新后的Excel文件
    workbook.save(file_path)


def get_first_column_values_as_set(file_path, sheet_name):
    # 打开指定的xlsx文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取指定的工作表
    sheet = workbook[sheet_name]
    # 获取第一列的最大行号
    max_row = sheet.max_row
    # 第一列的字母表示
    first_column_letter = get_column_letter(1)

    # 创建一个集合来存储第一列的所有单元格内容
    column_values_set = set()

    # 遍历第一列的每个单元格，并将其值添加到集合中
    for row in range(1, max_row + 1):
        cell_value = sheet[f"{first_column_letter}{row}"].value
        column_values_set.add(cell_value)

    return column_values_set
